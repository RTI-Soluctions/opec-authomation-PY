import os
import openpyxl
from datetime import datetime

campaign1 = "PI 70485"

pathLogFiles = "src/logs/"

actual_date = datetime.now()

date_format = "%H:%M:%S - %d/%m/%Y"

hourFormat = "%H:%M:%S"

organized_date = actual_date.strftime(date_format)


def campaignExtractor(campaign):
    exibitionList = []

    for file_name in os.listdir(pathLogFiles):
        fileLog = os.path.join(pathLogFiles, file_name)

        if fileLog.endswith(".log"):
            with open(fileLog, "r", encoding="utf-16-le") as file:
                linhas = file.readlines()

            lineList = [word.strip("\n") for word in linhas]

            for line in lineList:
                if campaign in line:
                    newLine = line.split("\t")
                    dateHourObject = datetime.strptime(newLine[0], "%d-%m-%Y %H:%M:%S")
                    if dateHourObject.hour == 7:
                        newLine.append("TJM1ED")
                    elif dateHourObject.hour == 12 or dateHourObject.hour == 12:
                        newLine.append("TJM2ED")
                    elif dateHourObject.hour == 19 or dateHourObject.hour == 20:
                        newLine.append("TJM3ED")
                    else:
                        newLine.append("ROTATIVO")
                    exibitionList.append(newLine)

    final_file = "Report_Template.xlsx"

    cellsPosition = [1, 3, 5, 10, 11]

    def insertDataIntoCells(sheet, data, start_line):
        sheet.cell(row=8, column=11, value=organized_date)

        for i, line in enumerate(data):
            for j, value in enumerate(line):
                for k, position in enumerate(cellsPosition):
                    if j == k:
                        sheet.cell(row=start_line + i, column=position, value=value)

    workbook = openpyxl.load_workbook(final_file)

    sheet = workbook.worksheets[0]

    start_line = 20

    insertDataIntoCells(sheet, exibitionList, start_line)

    workbook.save(f"Report_Template.xlsx")

    print(f"Relatório {final_file} atualizado com sucesso.")


campaignExtractor(campaign1)

# window.mainloop()
